import openai
import os
from decouple import config
import openpyxl


def ReadtheInput():
    wrkbk = openpyxl.load_workbook("test.xlsx")

    sh = wrkbk.active

    # iterate through excel and display data
    mylist = []
    for i in range(1, sh.max_row + 1):
        # print("\n")
        # print("Row ", i, " data :")

        for j in range(1, sh.max_column + 1):
            cell_obj = sh.cell(row=i, column=j)
            # print(cell_obj.value, end=" ")
            mylist.append(cell_obj.value)

    return mylist


def chatwithme(value):
    key=os.environ["API_KEY"] = "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"
    openai.api_key = key
    model_engine = 'text-davinci-002'
    input1 = value
    prompt = "using Selenium 4.3 with java write code for following scenario:" + input1 + " using with dynamic wait at each step"
    print(prompt)

    completion = openai.Completion.create(
        engine=model_engine,
        prompt=prompt,
        max_tokens=1024,
        n=1,
        stop=None,
        temperature=0.5,
    )
    response = completion.choices[0].text
    print(response)
    return response


def createFile(data):
    fp = open('TestCase.java', 'w')
    fp.write(data)
    fp.close()


myval = ReadtheInput()
resp = chatwithme(str(myval))
createFile(resp)
